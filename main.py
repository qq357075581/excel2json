import codecs
import os

import json

from openpyxl import load_workbook


class ExcelUtils:
    def __init__(self):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.sheetnames
        self.sheet = sheets[0]
        # print(sheets[0])
        self.ws = self.wb[self.sheet]

    # 行数
    def get_rows(self):
        rows = self.ws.max_row
        return rows

    # 列数
    def get_clos(self):
        clo = self.ws.max_column
        return clo

    # 获取值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 修改值并保存
    def set_cell_value(self, row, column, cell_value):
        try:
            self.ws.cell(row=row, column=column).value = cell_value
            self.wb.save(self.file)
        except Exception as e:
            print("error :{}".format(e))
            self.wb.save(self.file)


def to_json():
    excel_utils = ExcelUtils()
    excel_list = []

    clo = excel_utils.get_clos()
    # 遍历excel中的值存入字典中
    for j in range(3, excel_utils.get_rows() + 1):
        excel_dict = {}
        for i in range(1, clo + 1):
            dict_key = excel_utils.get_cell_value(2, i)
            dict_value = excel_utils.get_cell_value(j, i)
            if dict_key:
                excel_dict[dict_key] = str(dict_value)
            continue
        excel_list.append(excel_dict)

    data = {
        'items': excel_list,
    }
    # 字典转json
    # excel_json = json.dumps(excel_list)
    excel_json = json.dumps(data, indent=1, ensure_ascii=False)
    writeJson(fileName, excel_json)
    return excel_json


def writeJson(name, json):
    with open(jsonPath + name + ".json", 'w+', encoding="utf-8") as fp:
        fp.write(json)


# store  achievement
fileName = "achievement"
file = "./excel/" + fileName + ".xlsx"
jsonPath = "./json/"
un_use_str = ''

def file_name(file_dir):
    for root, dirs, files in os.walk(file_dir):
        # print(root)  # 当前目录路径
        for f in files:
            px = os.path.splitext(f)[1]
            if px == '.xlsx':
                global fileName, file, jsonPath
                fileName = os.path.splitext(f)[0]
                file = root + "/" + f
                to_json()
                print("finish=>",f)
            else:
                print(f,"文件格式不正确，仅支持.xlsx格式")


if __name__ == "__main__":
    file_name("./excel")
